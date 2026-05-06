from django.shortcuts import get_object_or_404, render, redirect
from django.http import Http404
from django.urls import reverse
from django.db.models import Q
from datetime import timedelta
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.db import models
from .models import Evento, Momento, Foto, Asistente, normalizar_texto
from .forms import FotoUploadForm
from usuarios.views import login_view
from usuarios.urls import urlpatterns as usuarios_urlpatterns
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count
from openpyxl import load_workbook
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def intro(request, token):
    evento = get_object_or_404(Evento, token=token, activo=True)

    asistentes = []
    busqueda = request.GET.get("q", "").strip()
    mensaje = None
    error = None

    if busqueda:
        texto_busqueda = normalizar_texto(busqueda)
        palabras = texto_busqueda.split()

        qs = Asistente.objects.filter(evento=evento)

        for palabra in palabras:
            qs = qs.filter(nombre_busqueda__icontains=palabra)

        asistentes = qs[:15]

    if request.method == "POST":
        asistente_id = request.POST.get("asistente_id")
        accion = request.POST.get("accion")

        asistente = get_object_or_404(Asistente, id=asistente_id, evento=evento)

        if accion == "no_asiste":
            asistente.estado_confirmacion = Asistente.ESTADO_NO_ASISTE
            asistente.adultos_confirmados = 0
            asistente.ninos_confirmados = 0
            asistente.confirmado_en = timezone.now()
            asistente.save()

            mensaje = "Respuesta registrada correctamente. Gracias por confirmar."

        else:
            adultos_confirmados = request.POST.get("adultos_confirmados")
            ninos_confirmados = request.POST.get("ninos_confirmados")

            try:
                adultos_confirmados = int(adultos_confirmados)
            except:
                adultos_confirmados = 0

            try:
                ninos_confirmados = int(ninos_confirmados)
            except:
                ninos_confirmados = 0

            if adultos_confirmados < 0 or ninos_confirmados < 0:
                error = "Los valores confirmados no pueden ser negativos."
            elif adultos_confirmados > asistente.adultos:
                error = f"No puede confirmar mas de {asistente.adultos} adulto(s)."
            elif ninos_confirmados > asistente.ninos:
                error = f"No puede confirmar mas de {asistente.ninos} nino(s)."
            elif adultos_confirmados + ninos_confirmados < 1:
                error = "Debe confirmar al menos 1 asistente o seleccionar No asistire."
            else:
                asistente.estado_confirmacion = Asistente.ESTADO_ASISTE
                asistente.adultos_confirmados = adultos_confirmados
                asistente.ninos_confirmados = ninos_confirmados
                asistente.confirmado_en = timezone.now()
                asistente.save()

                mensaje = "Asistencia confirmada correctamente. Gracias."

    return render(request, "memories/intro.html", {
        "evento": evento,
        "asistentes": asistentes,
        "busqueda": busqueda,
        "mensaje": mensaje,
        "error": error,
    })

def _get_anon_session_id(request):
    anon_id = request.session.get("anon_id")
    if not anon_id:
        request.session["anon_id"] = timezone.now().strftime("%Y%m%d%H%M%S%f")
        request.session.modified = True
        anon_id = request.session["anon_id"]
    return anon_id

def evento_home(request, token):
    evento = get_object_or_404(Evento, token=token, activo=True)
    momentos = Momento.objects.filter(evento=evento, activo=True).order_by("orden")

    return render(request, "memories/evento_home.html", {
        "evento": evento,
        "momentos": momentos,
    })




def subir_fotos(request, token):
    evento = get_object_or_404(Evento, token=token, activo=True)

    if request.method == "POST":
        form = FotoUploadForm(request.POST, request.FILES, evento=evento)
        if form.is_valid():
            momento = form.cleaned_data["momento"]
            nombre_invitado = (form.cleaned_data.get("nombre_invitado") or "").strip()
            mensaje = (form.cleaned_data.get("mensaje") or "").strip()
            
            anon_id = _get_anon_session_id(request)
            imagenes = form.cleaned_data["imagenes"]
            for img in imagenes:
                Foto.objects.create(
                    evento=evento,
                    momento=momento,
                    imagen=img,
                    nombre_invitado=nombre_invitado,
                    mensaje=mensaje,
                    owner_session_id=anon_id,
                    visible=True,
                )

            return redirect("memories:muro", token=evento.token)
    else:
        # ✅ Preselección SOLO si viene ?momento=<id>
        initial = {}
        momento_id = request.GET.get("momento")
        if momento_id:
            try:
                initial["momento"] = Momento.objects.get(id=momento_id, evento=evento, activo=True)
            except Momento.DoesNotExist:
                pass

        form = FotoUploadForm(evento=evento, initial=initial)

    return render(request, "memories/subir_fotos.html", {
        "evento": evento,
        "form": form,
    })


def muro(request, token):
    evento = get_object_or_404(Evento, token=token, activo=True)

    momento_id = request.GET.get("momento")
    momentos = Momento.objects.filter(evento=evento, activo=True).order_by("orden")

    fotos = Foto.objects.filter(evento=evento, visible=True).select_related("momento")

    momento_seleccionado = None
    if momento_id:
        try:
            momento_seleccionado = momentos.get(id=momento_id)
            fotos = fotos.filter(momento=momento_seleccionado)
        except Momento.DoesNotExist:
            momento_seleccionado = None

    # Limitar carga inicial (mejor para 200 invitados)
    fotos = fotos.order_by("-creado_en")[:240]
    anon_id = _get_anon_session_id(request)
    limite = timezone.now() - timedelta(minutes=10)

    # Marcamos flags en cada foto (para la plantilla)
    for f in fotos:
        f.es_mia = (f.owner_session_id == anon_id) and bool(f.owner_session_id)
        f.puede_ocultar = f.es_mia and (f.creado_en >= limite)

    return render(request, "memories/muro.html", {
        "evento": evento,
        "momentos": momentos,
        "fotos": fotos,
        "momento_seleccionado": momento_seleccionado,
    })



def foto_detalle(request, token, foto_id):
    evento = get_object_or_404(Evento, token=token, activo=True)
    foto = get_object_or_404(
        Foto.objects.select_related("momento"),
        id=foto_id,
        evento=evento,
        visible=True,
    )

    momento_id = request.GET.get("momento")
    back_url = reverse("memories:muro", kwargs={"token": token})
    if momento_id:
        back_url = f"{back_url}?momento={momento_id}"

    return render(request, "memories/foto_detalle.html", {
        "evento": evento,
        "foto": foto,
        "back_url": back_url,
    })


@require_POST
def ocultar_foto(request, token, foto_id):
    evento = get_object_or_404(Evento, token=token, activo=True)
    foto = get_object_or_404(Foto, id=foto_id, evento=evento)

    # ✅ Si está logueado, puede ocultar cualquier foto
    if request.user.is_authenticated:
        foto.visible = False
        foto.save(update_fields=["visible"])

    else:
        # ✅ Si NO está logueado: solo su sesión + solo 10 min
        anon_id = _get_anon_session_id(request)
        limite = timezone.now() - timedelta(minutes=10)

        if foto.owner_session_id != anon_id:
            raise Http404("No autorizado")

        if foto.creado_en < limite:
            raise Http404("Tiempo expirado")

        foto.visible = False
        foto.save(update_fields=["visible"])

    # Volver al muro conservando filtro si viene
    momento = request.POST.get("momento")
    if momento:
        return redirect(f"{reverse('memories:muro', kwargs={'token': token})}?momento={momento}")

    return redirect("memories:muro", token=token)

@login_required
def asistentes_admin(request, token):
    evento = get_object_or_404(Evento, token=token, activo=True)

    if request.method == "POST":
        archivo = request.FILES.get("archivo_excel")

        if not archivo:
            messages.error(request, "Debe seleccionar un archivo Excel.")
            return redirect("memories:asistentes_admin", token=evento.token)

        try:
            wb = load_workbook(archivo, data_only=True)
            ws = wb.active

            creados = 0
            actualizados = 0
            omitidos = 0

            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value).strip().lower() if cell.value else "")

            def obtener(row, nombre_columna):
                try:
                    index = headers.index(nombre_columna)
                    value = row[index]
                    return str(value).strip() if value is not None else ""
                except ValueError:
                    return ""

            for row in ws.iter_rows(min_row=2, values_only=True):
                nombres = obtener(row, "nombres")
                adultos = obtener(row, "adultos")
                ninos = obtener(row, "ninos")

                if not nombres:
                    omitidos += 1
                    continue

                try:
                    adultos = int(float(adultos))
                except:
                    adultos = 0

                try:
                    ninos = int(float(ninos))
                except:
                    ninos = 0

                if adultos < 0:
                    adultos = 0

                if ninos < 0:
                    ninos = 0

                if adultos + ninos < 1:
                    omitidos += 1
                    continue

                existente = Asistente.objects.filter(
                    evento=evento,
                    nombres__iexact=nombres,
                ).first()

                if existente:
                    existente.adultos = adultos
                    existente.ninos = ninos
                    existente.save()
                    actualizados += 1
                else:
                    Asistente.objects.create(
                        evento=evento,
                        nombres=nombres,
                        adultos=adultos,
                        ninos=ninos,
                    )
                    creados += 1

            messages.success(
                request,
                f"Importacion finalizada. Creados: {creados}, actualizados: {actualizados}, omitidos: {omitidos}."
            )

        except Exception as e:
            messages.error(request, f"Error al procesar el Excel: {e}")

        return redirect("memories:asistentes_admin", token=evento.token)

    asistentes = Asistente.objects.filter(evento=evento).order_by("nombres")

    total_asistentes = asistentes.count()
    confirmados = asistentes.filter(estado_confirmacion=Asistente.ESTADO_ASISTE).count()
    no_asisten = asistentes.filter(estado_confirmacion=Asistente.ESTADO_NO_ASISTE).count()
    pendientes = asistentes.filter(estado_confirmacion=Asistente.ESTADO_PENDIENTE).count()

    total_adultos = asistentes.aggregate(total=Sum("adultos"))["total"] or 0
    total_ninos = asistentes.aggregate(total=Sum("ninos"))["total"] or 0

    total_adultos_confirmados = asistentes.aggregate(total=Sum("adultos_confirmados"))["total"] or 0
    total_ninos_confirmados = asistentes.aggregate(total=Sum("ninos_confirmados"))["total"] or 0

    return render(request, "memories/asistentes_admin.html", {
        "evento": evento,
        "asistentes": asistentes,
        "total_asistentes": total_asistentes,
        "confirmados": confirmados,
        "no_asisten": no_asisten,
        "pendientes": pendientes,
        "total_adultos": total_adultos,
        "total_ninos": total_ninos,
        "total_adultos_confirmados": total_adultos_confirmados,
        "total_ninos_confirmados": total_ninos_confirmados,
    })

@login_required
def exportar_asistentes_excel(request, token):
    evento = get_object_or_404(Evento, token=token, activo=True)
    asistentes = Asistente.objects.filter(evento=evento).order_by("nombres")

    wb = Workbook()
    ws = wb.active
    ws.title = "Confirmaciones"

    # Colores
    azul = "0E1424"
    dorado = "CC9933"
    verde = "D9EAD3"
    rojo = "F4CCCC"
    amarillo = "FFF2CC"
    gris = "F3F4F6"

    fill_title = PatternFill("solid", fgColor=azul)
    fill_header = PatternFill("solid", fgColor=dorado)
    fill_success = PatternFill("solid", fgColor=verde)
    fill_danger = PatternFill("solid", fgColor=rojo)
    fill_pending = PatternFill("solid", fgColor=amarillo)
    fill_gray = PatternFill("solid", fgColor=gris)

    font_title = Font(color="FFFFFF", bold=True, size=14)
    font_header = Font(color="000000", bold=True)
    font_bold = Font(bold=True)

    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Titulo
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Resumen de confirmaciones - {evento.nombre}"
    ws["A1"].fill = fill_title
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="center")

    # Totales
    total_registros = asistentes.count()
    total_asisten = asistentes.filter(estado_confirmacion=Asistente.ESTADO_ASISTE).count()
    total_no_asisten = asistentes.filter(estado_confirmacion=Asistente.ESTADO_NO_ASISTE).count()
    total_pendientes = asistentes.filter(estado_confirmacion=Asistente.ESTADO_PENDIENTE).count()

    total_adultos = sum(a.adultos for a in asistentes)
    total_ninos = sum(a.ninos for a in asistentes)
    total_adultos_confirmados = sum(a.adultos_confirmados for a in asistentes)
    total_ninos_confirmados = sum(a.ninos_confirmados for a in asistentes)

    resumen = [
        ("Total registros", total_registros),
        ("Si asistiran", total_asisten),
        ("No asistiran", total_no_asisten),
        ("Pendientes", total_pendientes),
        ("Adultos confirmados", f"{total_adultos_confirmados} / {total_adultos}"),
        ("Ninos confirmados", f"{total_ninos_confirmados} / {total_ninos}"),
    ]

    row = 3
    for etiqueta, valor in resumen:
        ws[f"A{row}"] = etiqueta
        ws[f"B{row}"] = valor
        ws[f"A{row}"].font = font_bold
        ws[f"A{row}"].fill = fill_gray
        ws[f"A{row}"].border = border
        ws[f"B{row}"].border = border
        row += 1

    # Encabezados
    headers = [
        "#",
        "Nombres",
        "Adultos asignados",
        "Ninos asignados",
        "Total pases",
        "Estado",
        "Adultos confirmados",
        "Ninos confirmados",
        "Total confirmados",
    ]

    start_row = 11

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col)
        cell.value = header
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Datos
    for idx, a in enumerate(asistentes, start=1):
        fila = start_row + idx

        if a.estado_confirmacion == Asistente.ESTADO_ASISTE:
            estado = "Asiste"
            fill_estado = fill_success
        elif a.estado_confirmacion == Asistente.ESTADO_NO_ASISTE:
            estado = "No asiste"
            fill_estado = fill_danger
        else:
            estado = "Pendiente"
            fill_estado = fill_pending

        datos = [
            idx,
            a.nombres,
            a.adultos,
            a.ninos,
            a.total_pases,
            estado,
            a.adultos_confirmados,
            a.ninos_confirmados,
            a.total_confirmados,
        ]

        for col, valor in enumerate(datos, start=1):
            cell = ws.cell(row=fila, column=col)
            cell.value = valor
            cell.border = border
            cell.alignment = Alignment(vertical="center")

            if col == 6:
                cell.fill = fill_estado
                cell.font = font_bold

    # Ajustes visuales
    widths = {
        "A": 6,
        "B": 38,
        "C": 18,
        "D": 16,
        "E": 14,
        "F": 16,
        "G": 22,
        "H": 20,
        "I": 20,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A12"
    ws.auto_filter.ref = f"A{start_row}:I{start_row + total_registros}"

    filename = f"confirmaciones_{evento.token}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response